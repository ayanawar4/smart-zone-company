import openpyxl, json, datetime, re

SRC = "../seed-data-source.xlsx"
OUT = "../seed-data"

wb = openpyxl.load_workbook(SRC, data_only=True)

def to_date(v):
    if v is None:
        return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if not s:
        return None
    # try dd/mm/yyyy
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', s)
    if m:
        d, mo, y = m.groups()
        try:
            return f"{y}-{int(mo):02d}-{int(d):02d}"
        except Exception:
            return s
    return s

def num(v):
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    try:
        return round(float(str(v).replace(',', '').strip()), 2)
    except Exception:
        return 0

PROJECT_SHEETS = ['N.C','SERVER','RTR','HILTON K.R','A.P','SHOMAN HOME','OROUBA CAM',
                  'MR MAMDOUH','MR MAMDOUH (2)','HILTON CORNECH','SPA VILLA','HILTON CAM',
                  'HILTON CORNECH 2','HILTON G.P 300s']

projects = []
for name in PROJECT_SHEETS:
    ws = wb[name]
    cost_items = []
    for r in range(3, 42):
        item = ws.cell(r, 15).value   # O
        if item is None or str(item).strip() == '':
            continue
        cost_items.append({
            "no": ws.cell(r, 14).value,
            "item": str(item).strip(),
            "unitPrice": num(ws.cell(r, 16).value),   # P
            "qty": num(ws.cell(r, 17).value),          # Q
            "totalPrice": num(ws.cell(r, 18).value),   # R
        })
    install_items = []
    for r in range(3, 48):
        nm = ws.cell(r, 22).value  # V
        if nm is None or str(nm).strip() == '':
            continue
        install_items.append({
            "no": ws.cell(r, 20).value,          # T
            "date": to_date(ws.cell(r, 21).value),  # U
            "name": str(nm).strip(),
            "unitPrice": num(ws.cell(r, 23).value),  # W
            "qty": num(ws.cell(r, 24).value),        # X
            "totalPrice": num(ws.cell(r, 25).value), # Y
        })
    fund_history = []
    for r in range(3, 37):
        v = ws.cell(r, 5).value  # E
        if v is None or not isinstance(v, (int, float)):
            continue
        fund_history.append(num(v))

    summary = {
        "projectCostTotal": num(ws.cell(12, 11).value),   # K12
        "totalPayed": num(ws.cell(15, 7).value),           # G15
        "fund": num(ws.cell(15, 8).value),                 # H15
        "hesham": num(ws.cell(15, 9).value),                # I15
        "sayed": num(ws.cell(15, 10).value),                # J15
        "projectAmount": num(ws.cell(15, 11).value),        # K15
        "remain": num(ws.cell(15, 12).value),                # L15
        "installationTotal": num(ws.cell(47, 25).value) if ws.cell(47,21).value == 'total' else None,
        "fundColumnTotal": num(ws.cell(37, 5).value),  # E37
        "heshamColumnTotal": num(ws.cell(37, 2).value),  # B37
        "sayedColumnTotal": num(ws.cell(37, 3).value),  # C37
        "projectColumnTotal": num(ws.cell(37, 4).value),  # D37
    }

    projects.append({
        "name": name.strip(),
        "costItems": cost_items,
        "installItems": install_items,
        "fundHistory": fund_history,
        "summary": summary,
    })

with open(f"{OUT}/projects.json", "w", encoding="utf-8") as f:
    json.dump(projects, f, ensure_ascii=False, indent=2)
print("projects:", len(projects), "sample cost items p0:", len(projects[0]['costItems']))

# ---------------- FUND LEDGER ----------------
ws = wb['fund ']
fund_tx = []
for r in range(3, ws.max_row + 1):
    no = ws.cell(r, 4).value  # D
    item = ws.cell(r, 6).value  # F
    if no is None and item is None:
        continue
    fund_tx.append({
        "no": no,
        "project": (str(ws.cell(r,5).value).strip() if ws.cell(r,5).value else None),  # E
        "item": (str(item).strip() if item else None),
        "category": (str(ws.cell(r,7).value).strip() if ws.cell(r,7).value else None),  # G
        "amountIn": num(ws.cell(r, 8).value),   # H
        "amountOut": num(ws.cell(r, 9).value),  # I
        "balanceAfter": num(ws.cell(r, 10).value),  # J
        "note": (str(ws.cell(r,11).value).strip() if ws.cell(r,11).value else None),  # K
    })
with open(f"{OUT}/fund_transactions.json", "w", encoding="utf-8") as f:
    json.dump(fund_tx, f, ensure_ascii=False, indent=2)
print("fund tx:", len(fund_tx))

# ---------------- EXPENSES LEDGER (المصاريف: source/amount/date cols I,J,K) ----------------
ws = wb['المصاريف']
expenses = []
for r in range(4, 67):
    src = ws.cell(r, 9).value   # I
    amt = ws.cell(r, 10).value  # J
    if src is None and amt is None:
        continue
    if str(src).strip().upper() == 'TOTAL':
        continue
    expenses.append({
        "source": (str(src).strip() if src else None),
        "amount": num(amt),
        "date": to_date(ws.cell(r, 11).value),  # K
        "note": (str(ws.cell(r,12).value).strip() if ws.cell(r,12).value else None),  # L
    })
with open(f"{OUT}/expenses.json", "w", encoding="utf-8") as f:
    json.dump(expenses, f, ensure_ascii=False, indent=2)
print("expenses:", len(expenses))

# ---------------- SALARIES (wide table melt) ----------------
# Block 1: columns P..AA = months 12..1, name in AB, rows 4-9ish (year ~2025 first half of table)
# Block 2: columns AG..AR = months 12..1, name in AS, rows 4-11ish
salaries = []
def melt_block(name_col, month_cols, year_label, row_start=4, row_end=11):
    for r in range(row_start, row_end + 1):
        emp = ws.cell(r, name_col).value
        if emp is None or str(emp).strip() == '' or str(emp).strip() in ('الاجمالى',):
            continue
        emp_name = str(emp).strip()
        for month_num, col in month_cols:
            v = ws.cell(r, col).value
            if v is None:
                continue
            if isinstance(v, (int, float)):
                salaries.append({
                    "employeeName": emp_name,
                    "month": month_num,
                    "year": year_label,
                    "amount": num(v),
                    "note": None,
                })
            else:
                # text like "مرتب+حافز" or "1000+8000" -> try to sum numeric parts, else store as note
                s = str(v).strip()
                parts = re.findall(r'-?\d+(?:\.\d+)?', s)
                if parts and re.fullmatch(r'[\d\.\+\-\s]+', s):
                    salaries.append({
                        "employeeName": emp_name, "month": month_num, "year": year_label,
                        "amount": sum(float(p) for p in parts), "note": s,
                    })
                else:
                    salaries.append({
                        "employeeName": emp_name, "month": month_num, "year": year_label,
                        "amount": 0, "note": s,
                    })

months_block1 = [(12,16),(11,17),(10,18),(9,19),(8,20),(7,21),(6,22),(5,23),(4,24),(3,25),(2,26),(1,27)]  # P..AA
melt_block(28, months_block1, 2025, 4, 9)  # AB

months_block2 = [(12,33),(11,34),(10,35),(9,36),(8,37),(7,38),(6,39),(5,40),(4,41),(3,42),(2,43),(1,44)]  # AG..AR
melt_block(45, months_block2, 2026, 4, 11)  # AS

with open(f"{OUT}/salaries.json", "w", encoding="utf-8") as f:
    json.dump(salaries, f, ensure_ascii=False, indent=2)
print("salaries:", len(salaries))

# ---------------- INVOICES (smart zone) ----------------
ws = wb['smart zone ']
invoices = []
for r in range(4, ws.max_row + 1):
    no = ws.cell(r, 4).value  # D
    if no is None:
        continue
    invoices.append({
        "no": no,
        "date": to_date(ws.cell(r, 5).value),        # E
        "invoiceNo": ws.cell(r, 6).value,             # F
        "customer": (str(ws.cell(r,7).value).strip() if ws.cell(r,7).value else None),  # G
        "project": (str(ws.cell(r,8).value).strip() if ws.cell(r,8).value else None),   # H
        "deposit": num(ws.cell(r, 9).value),          # I
        "subtotal": num(ws.cell(r, 10).value),        # J
        "vat": num(ws.cell(r, 11).value),              # K
        "total": num(ws.cell(r, 12).value),            # L
        "statusDone": bool(ws.cell(r,13).value and str(ws.cell(r,13).value).strip()),   # M
        "statusUnderConstruction": bool(ws.cell(r,14).value and str(ws.cell(r,14).value).strip()),  # N
        "commission3pct": num(ws.cell(r, 15).value),   # O
        "commission1pct": num(ws.cell(r, 16).value),   # P
        "remainSmartZone": num(ws.cell(r, 17).value),  # Q
        "remainCustomer": num(ws.cell(r, 18).value),   # R
    })
with open(f"{OUT}/invoices.json", "w", encoding="utf-8") as f:
    json.dump(invoices, f, ensure_ascii=False, indent=2)
print("invoices:", len(invoices))

# ---------------- Employees derived ----------------
names = set()
for s in salaries:
    names.add(s['employeeName'])
for name in ['HESHAM','SAYED']:
    names.add(name)
employees = [{"name": n, "role": None, "active": True} for n in sorted(names)]
with open(f"{OUT}/employees.json", "w", encoding="utf-8") as f:
    json.dump(employees, f, ensure_ascii=False, indent=2)
print("employees:", len(employees))
